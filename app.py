import streamlit as st
import pandas as pd
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

st.title("📦 VCC Warehouse Packing Group Generator")

uploaded_file = st.file_uploader("Upload your POInput file (CSV or Excel)", type=["xlsx", "csv"])

if uploaded_file is not None:

    if uploaded_file.name.endswith('.csv'):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    required_cols = {'PO Number', 'Material Description', 'Style Code', 'Size', 'Article Qty'}
    if not required_cols.issubset(set(df.columns)):
        st.error(f"❌ Uploaded file is missing required columns. Please include: {', '.join(required_cols)}")
        st.stop()

    def extract_color(desc):
        if pd.isna(desc):
            return 'UNKNOWN'
        parts = desc.split(',')
        for part in parts:
            cleaned = part.strip().upper().replace('.', '')
            if cleaned and all(c.isalpha() or c.isspace() for c in cleaned) and len(cleaned) >= 3:
                return cleaned
        return 'UNKNOWN'

    def extract_style_digits(style):
        if pd.isna(style):
            return ''
        match = re.search(r'(\d+)$', style)
        return match.group(1) if match else ''

    df['Color'] = df['Material Description'].apply(extract_color)
    df['StyleDigits'] = df['Style Code'].apply(extract_style_digits)
    df['ColorStyle'] = df['Color'] + ' - ' + df['StyleDigits']

    # Update infant_sizes to include 0-3M and 3-6M
    infant_sizes = ['0-3M', '3-6M', '6-12M', '12-18M', '18-24M']
    toddler_sizes = ['2-3Y', '3-4Y', '5-6Y', '7-8Y']
    all_sizes = infant_sizes + toddler_sizes

    pivot = df.pivot_table(
        index=['PO Number', 'ColorStyle'],
        columns='Size',
        values='Article Qty',
        aggfunc='sum',
        fill_value=0
    )

    for col in all_sizes:
        if col not in pivot.columns:
            pivot[col] = 0

    pivot['Infant Total'] = pivot[infant_sizes].sum(axis=1)
    pivot['Toddler Total'] = pivot[toddler_sizes].sum(axis=1)
    pivot.reset_index(inplace=True)

    po_group_map = defaultdict(list)
    for po_number, group in pivot.groupby('PO Number'):
        signature = tuple(sorted(
            tuple([row['ColorStyle']] + [int(row.get(col, 0)) for col in all_sizes])
            for _, row in group.iterrows()
        ))
        po_group_map[signature].append(po_number)

    grouped_rows = []
    for idx, (sig, po_list) in enumerate(sorted(po_group_map.items()), start=1):
        po_count = len(set(po_list))
        for entry in sig:
            color_style = entry[0]
            size_vals = entry[1:]
            infant_values = size_vals[:len(infant_sizes)]
            toddler_values = size_vals[len(infant_sizes):]
            grouped_rows.append({
                'Group ID': f'Group {idx}',
                'ColorStyle': color_style,
                **{infant_sizes[i]: infant_values[i] for i in range(len(infant_sizes))},
                **{toddler_sizes[i]: toddler_values[i] for i in range(len(toddler_sizes))},
                'Infant Total': sum(infant_values),
                'Toddler Total': sum(toddler_values),
                'POs': ', '.join(map(str, sorted(set(po_list)))),
                'PO Count': po_count
            })

    grouped_df = pd.DataFrame(grouped_rows)
    grouped_df_sorted = grouped_df.sort_values(by='Group ID')
    group_ids = sorted(grouped_df_sorted['Group ID'].unique(), key=lambda g: int(g.replace('Group ', '')))

    wb = Workbook()
    ws = wb.active
    ws.title = "Packing Groups"
    current_row = 1

    for group_id in group_ids:
        group_data = grouped_df_sorted[grouped_df_sorted['Group ID'] == group_id]
        po_list = group_data['POs'].iloc[0]
        po_count = group_data['PO Count'].iloc[0]

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        header_cell = ws.cell(row=current_row, column=1, value=f"{group_id} (PO Count: {po_count})")
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center')
        current_row += 1

        # Infant sizes
        ws.cell(row=current_row, column=1, value="ColorStyle")
        for i, size in enumerate(infant_sizes, start=2):
            ws.cell(row=current_row, column=i, value=size)
        ws.cell(row=current_row, column=i+1, value="Total")
        current_row += 1

        for _, row in group_data.iterrows():
            if row['Infant Total'] > 0:
                ws.cell(row=current_row, column=1, value=row['ColorStyle'])
                for i, size in enumerate(infant_sizes, start=2):
                    ws.cell(row=current_row, column=i, value=row.get(size, 0))
                ws.cell(row=current_row, column=i+1, value=row['Infant Total'])
                current_row += 1

        current_row += 1

        # Toddler sizes
        ws.cell(row=current_row, column=1, value="ColorStyle")
        for i, size in enumerate(toddler_sizes, start=2):
            ws.cell(row=current_row, column=i, value=size)
        ws.cell(row=current_row, column=i+1, value="Total")
        current_row += 1

        for _, row in group_data.iterrows():
            if row['Toddler Total'] > 0:
                ws.cell(row=current_row, column=1, value=row['ColorStyle'])
                for i, size in enumerate(toddler_sizes, start=2):
                    ws.cell(row=current_row, column=i, value=row.get(size, 0))
                ws.cell(row=current_row, column=i+1, value=row['Toddler Total'])
                current_row += 1

        # Associated POs
        current_row += 1
        ws.cell(row=current_row, column=1, value="Associated POs:")
        po_list_split = po_list.split(', ')
        for i, po in enumerate(po_list_split):
            ws.cell(row=current_row + i, column=2, value=po)
        current_row += len(po_list_split)
        current_row += 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success("✅ Report generated!")
    st.download_button("📥 Download Packing Report", output, file_name="Packing_Group_Report.xlsx")
